import json
import uuid
from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.transaction import Transaction, TransactionType
from app.models.user import User
from app.services.portfolio.positions import calculate_positions

router = APIRouter(prefix="/portfolio", tags=["export"])

_BRAND      = colors.HexColor("#1D4ED8")
_BRAND_DARK = colors.HexColor("#1E40AF")
_BRAND_LIGHT = colors.HexColor("#EFF6FF")
_ALT_ROW    = colors.HexColor("#F8FAFC")
_GRAY       = colors.HexColor("#6B7280")

_HDR_FILL = PatternFill("solid", fgColor="1D4ED8")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="EFF6FF")


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18


def _autowidth(ws, max_col_width: int = 42) -> None:
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, max_col_width)


def _zebra(ws, start_row: int = 2) -> None:
    for i, row in enumerate(ws.iter_rows(min_row=start_row)):
        if i % 2 == 0:
            for cell in row:
                cell.fill = _ALT_FILL


@router.get("/export")
async def export_portfolio(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Esporta posizioni e transazioni in formato Excel (.xlsx)."""
    stmt = (
        select(Transaction)
        .join(Transaction.account)
        .where(Transaction.account.has(user_id=current_user.id))
        .options(
            selectinload(Transaction.asset),
            selectinload(Transaction.account),
        )
        .order_by(Transaction.date.asc(), Transaction.id.asc())
    )
    result = await db.execute(stmt)
    transactions = result.scalars().all()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Transazioni ─────────────────────────────────────────────────
    ws_tx = wb.active
    ws_tx.title = "Transazioni"
    ws_tx.append([
        "Data", "Tipo", "Simbolo", "Nome asset", "ISIN",
        "Quantità", "Prezzo", "Valuta", "Tasso cambio",
        "Commissioni", "Totale EUR", "Conto", "Note",
    ])
    _style_header(ws_tx)

    asset_map: dict[int, object] = {}
    for tx in transactions:
        asset_map[tx.asset_id] = tx.asset
        total_eur = round(tx.quantity * tx.price * tx.exchange_rate + tx.fee, 2)
        ws_tx.append([
            tx.date.isoformat(),
            tx.type.value if hasattr(tx.type, "value") else str(tx.type),
            tx.asset.symbol,
            tx.asset.name,
            tx.asset.isin or "",
            round(tx.quantity, 6),
            round(tx.price, 6),
            tx.price_currency,
            round(tx.exchange_rate, 6),
            round(tx.fee, 2),
            total_eur,
            tx.account.name,
            tx.notes or "",
        ])

    _zebra(ws_tx)
    _autowidth(ws_tx)

    # ── Sheet 2: Posizioni aperte ────────────────────────────────────────────
    ws_pos = wb.create_sheet("Posizioni")
    ws_pos.append([
        "Simbolo", "Nome", "Tipo", "ISIN", "Borsa",
        "Quantità", "PMC (EUR)", "Totale investito (EUR)", "P&L realizzato (EUR)",
    ])
    _style_header(ws_pos)

    positions = calculate_positions(list(transactions))
    for asset_id, pos in sorted(positions.items(), key=lambda kv: kv[0]):
        asset = asset_map.get(asset_id)
        if not asset:
            continue
        ws_pos.append([
            asset.symbol,
            asset.name,
            asset.type.value if hasattr(asset.type, "value") else str(asset.type),
            asset.isin or "",
            asset.exchange.value if hasattr(asset.exchange, "value") else str(asset.exchange),
            round(pos.quantity, 6),
            round(pos.pmc_eur, 4),
            round(pos.total_invested_eur, 2),
            round(pos.realized_pnl_eur, 2),
        ])

    _zebra(ws_pos)
    _autowidth(ws_pos)

    # ── Sheet 3: Info ────────────────────────────────────────────────────────
    ws_info = wb.create_sheet("Info")
    ws_info.append(["Campo", "Valore"])
    _style_header(ws_info)
    ws_info.append(["Data esportazione", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_info.append(["Utente", current_user.email])
    ws_info.append(["Transazioni totali", len(transactions)])
    ws_info.append(["Posizioni aperte", len(positions)])
    ws_info.column_dimensions["A"].width = 26
    ws_info.column_dimensions["B"].width = 32

    # Return as streaming xlsx
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"nextfolio_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/pdf")
async def export_pdf(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Esporta il portafoglio come report PDF (Posizioni + Transazioni)."""
    stmt = (
        select(Transaction)
        .join(Transaction.account)
        .where(Transaction.account.has(user_id=current_user.id))
        .options(
            selectinload(Transaction.asset),
            selectinload(Transaction.account),
        )
        .order_by(Transaction.date.asc(), Transaction.id.asc())
    )
    result = await db.execute(stmt)
    transactions = result.scalars().all()

    all_positions = calculate_positions(list(transactions), include_closed=True)
    positions = {aid: pos for aid, pos in all_positions.items() if pos.quantity > 1e-6}
    asset_map: dict[int, object] = {tx.asset_id: tx.asset for tx in transactions}

    buf = BytesIO()

    styles = getSampleStyleSheet()
    title_st  = ParagraphStyle("nf_title",  parent=styles["Normal"], fontName="Helvetica-Bold",
                               fontSize=22, textColor=_BRAND, spaceAfter=1*mm)
    sub_st    = ParagraphStyle("nf_sub",    parent=styles["Normal"], fontName="Helvetica",
                               fontSize=9,  textColor=_GRAY,  spaceAfter=6*mm, alignment=TA_RIGHT)
    section_st = ParagraphStyle("nf_sec",   parent=styles["Normal"], fontName="Helvetica-Bold",
                                fontSize=11, textColor=_BRAND, spaceBefore=6*mm, spaceAfter=3*mm)
    note_st   = ParagraphStyle("nf_note",   parent=styles["Normal"], fontName="Helvetica-Oblique",
                                fontSize=8,  textColor=_GRAY)

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(_GRAY)
        w, h = A4
        canvas.drawString(15*mm, 10*mm, "Nextfolio — Report riservato")
        canvas.drawRightString(w - 15*mm, 10*mm, f"Pagina {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=18*mm, bottomMargin=20*mm,
    )

    elems = []

    # ── Intestazione ──────────────────────────────────────────────────────────
    elems.append(Paragraph("Nextfolio", title_st))
    elems.append(Paragraph(f"Report portafoglio &bull; {now_str}", sub_st))
    elems.append(HRFlowable(width="100%", color=_BRAND, thickness=1.5, spaceAfter=5*mm))

    # ── Riepilogo ─────────────────────────────────────────────────────────────
    open_positions = positions
    total_invested = sum(p.total_invested_eur for p in open_positions.values())
    total_realized = sum(p.realized_pnl_eur   for p in all_positions.values())

    elems.append(Paragraph("Riepilogo", section_st))
    summary_rows = [
        ["Posizioni aperte",          str(len(open_positions))],
        ["Capitale investito",         f"EUR {total_invested:>14,.2f}"],
        ["P&L realizzato totale",      f"EUR {total_realized:>+14,.2f}"],
        ["Transazioni totali",         str(len(transactions))],
        ["Data esportazione",          now_str],
    ]
    summary_tbl = Table(summary_rows, colWidths=[85*mm, 70*mm])
    summary_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",  (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE",  (0, 0), (-1, -1), 9),
        ("ALIGN",     (1, 0), (1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, _ALT_ROW]),
        ("GRID",      (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
    ]))
    elems.append(summary_tbl)

    # ── Posizioni aperte ──────────────────────────────────────────────────────
    elems.append(Paragraph("Posizioni aperte", section_st))
    if open_positions:
        pos_hdr = ["Simbolo", "Nome", "Tipo", "Quantita'", "PMC (EUR)", "Investito (EUR)", "P&L Real. (EUR)"]
        pos_rows = [pos_hdr]
        for aid, pos in sorted(open_positions.items(),
                               key=lambda kv: kv[1].total_invested_eur, reverse=True):
            asset = asset_map.get(aid)
            if not asset:
                continue
            atype = asset.type.value if hasattr(asset.type, "value") else str(asset.type)
            pos_rows.append([
                asset.symbol,
                (asset.name[:32] + "…") if len(asset.name) > 32 else asset.name,
                atype,
                f"{pos.quantity:,.4f}",
                f"{pos.pmc_eur:,.4f}",
                f"{pos.total_invested_eur:>12,.2f}",
                f"{pos.realized_pnl_eur:>+12,.2f}",
            ])
        pos_tbl = Table(pos_rows,
                        colWidths=[20*mm, 54*mm, 16*mm, 22*mm, 22*mm, 28*mm, 28*mm],
                        repeatRows=1)
        pos_tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), _BRAND),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 0), (-1, -1), 7.5),
            ("ALIGN",        (3, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _ALT_ROW]),
            ("GRID",         (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING",   (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ]))
        elems.append(pos_tbl)
    else:
        elems.append(Paragraph("Nessuna posizione aperta.", note_st))

    # ── Transazioni ───────────────────────────────────────────────────────────
    elems.append(Paragraph("Transazioni", section_st))
    tx_hdr = ["Data", "Tipo", "Simbolo", "Quantita'", "Prezzo", "Cambio", "Totale EUR", "Conto"]
    tx_rows = [tx_hdr]
    for tx in transactions:
        total_eur = tx.quantity * tx.price * tx.exchange_rate + tx.fee
        tx_rows.append([
            tx.date.strftime("%d/%m/%Y"),
            tx.type.value if hasattr(tx.type, "value") else str(tx.type),
            tx.asset.symbol,
            f"{tx.quantity:,.4f}",
            f"{tx.price:,.4f}",
            f"{tx.exchange_rate:,.4f}",
            f"{total_eur:>10,.2f}",
            (tx.account.name[:18] + "…") if len(tx.account.name) > 18 else tx.account.name,
        ])
    tx_tbl = Table(tx_rows,
                   colWidths=[19*mm, 17*mm, 20*mm, 21*mm, 20*mm, 18*mm, 24*mm, 31*mm],
                   repeatRows=1)
    tx_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), _BRAND_DARK),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 0), (-1, -1), 7),
        ("ALIGN",        (3, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _ALT_ROW]),
        ("GRID",         (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
    ]))
    elems.append(tx_tbl)

    doc.build(elems, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)

    filename = f"nextfolio_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/ghostfolio")
async def export_ghostfolio(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Esporta il portafoglio in formato Ghostfolio v3.x (JSON).
    Compatibile con Ghostfolio import → Activities.
    """
    from app.models.account import Account

    stmt = (
        select(Transaction)
        .join(Transaction.account)
        .where(Transaction.account.has(user_id=current_user.id))
        .options(
            selectinload(Transaction.asset),
            selectinload(Transaction.account),
        )
        .order_by(Transaction.date.asc(), Transaction.id.asc())
    )
    result = await db.execute(stmt)
    transactions = list(result.scalars())

    accounts_result = await db.execute(
        select(Account).where(Account.user_id == current_user.id)
    )
    accounts = {a.id: a for a in accounts_result.scalars()}

    # Mappa tipo transazione → Ghostfolio type
    TX_TYPE_MAP = {
        TransactionType.BUY: "BUY",
        TransactionType.SELL: "SELL",
        TransactionType.DIVIDEND: "DIVIDEND",
        TransactionType.COUPON: "DIVIDEND",
        TransactionType.INTEREST: "INTEREST",
        TransactionType.FEE: "FEE",
    }

    # Costruisci la lista activities
    activities = []
    asset_profiles: dict[str, dict] = {}
    gf_accounts: dict[int, dict] = {}

    for tx in transactions:
        asset = tx.asset
        acc = accounts.get(tx.account_id)
        if not asset or not acc:
            continue

        gf_type = TX_TYPE_MAP.get(tx.type)
        if gf_type is None:
            continue

        # Aggiungi account se non già presente
        if tx.account_id not in gf_accounts:
            acc_gf_id = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"nextfolio-account-{tx.account_id}"))
            gf_accounts[tx.account_id] = {
                "id": acc_gf_id,
                "name": acc.name,
                "currency": acc.currency or "EUR",
                "platformId": None,
            }

        acc_gf_id = gf_accounts[tx.account_id]["id"]

        # Aggiungi asset profile se non già presente
        sym = asset.symbol
        if sym not in asset_profiles:
            asset_profiles[sym] = {
                "symbol": sym,
                "name": asset.name,
                "currency": asset.currency,
                "comment": asset.isin or "",
                "assetClass": {
                    "STOCK": "EQUITY",
                    "ETF": "EQUITY",
                    "BOND": "FIXED_INCOME",
                    "CRYPTO": "CRYPTOCURRENCY",
                    "COMMODITY": "COMMODITY",
                    "REIT": "REAL_ESTATE",
                }.get(asset.type.value if hasattr(asset.type, "value") else str(asset.type), "EQUITY"),
                "assetSubClass": {
                    "ETF": "ETF",
                    "STOCK": "STOCK",
                    "BOND": "BOND",
                }.get(asset.type.value if hasattr(asset.type, "value") else str(asset.type), "STOCK"),
                "dataSource": "YAHOO",
                "marketData": [],
            }

        activities.append({
            "accountId": acc_gf_id,
            "comment": tx.notes or "",
            "currency": tx.price_currency or asset.currency,
            "dataSource": "YAHOO",
            "date": datetime.combine(tx.date, datetime.min.time()).replace(tzinfo=timezone.utc).isoformat(),
            "fee": tx.fee,
            "quantity": tx.quantity,
            "symbol": sym,
            "type": gf_type,
            "unitPrice": tx.price,
        })

    gf_export = {
        "meta": {
            "date": datetime.now(timezone.utc).isoformat(),
            "version": "v3",
        },
        "accounts": list(gf_accounts.values()),
        "activities": activities,
        "assetProfiles": list(asset_profiles.values()),
    }

    filename = f"nextfolio_ghostfolio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    content = json.dumps(gf_export, ensure_ascii=False, indent=2)
    return StreamingResponse(
        iter([content]),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/nextfolio")
async def export_nextfolio(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Esporta tutti i dati in formato Nextfolio JSON.
    Il file è asettico: non contiene email, user_id o altri dati personali.
    Gli ID interni sono sostituiti con riferimenti posizionali autosufficienti.
    """
    from app.models.account import Account

    accounts_result = await db.execute(
        select(Account).where(Account.user_id == current_user.id).order_by(Account.id)
    )
    accounts = list(accounts_result.scalars())

    stmt = (
        select(Transaction)
        .join(Transaction.account)
        .where(Transaction.account.has(user_id=current_user.id))
        .options(
            selectinload(Transaction.asset),
            selectinload(Transaction.account),
        )
        .order_by(Transaction.date.asc(), Transaction.id.asc())
    )
    result = await db.execute(stmt)
    transactions = list(result.scalars())

    # Raccoglie gli asset unici nell'ordine in cui compaiono
    assets_seen: dict[int, object] = {}
    for tx in transactions:
        if tx.asset_id not in assets_seen:
            assets_seen[tx.asset_id] = tx.asset

    # Mappa DB id → id posizionale nel file (1-based)
    account_id_map = {a.id: idx + 1 for idx, a in enumerate(accounts)}
    asset_id_map = {a_id: idx + 1 for idx, a_id in enumerate(assets_seen)}

    def _val(obj, attr: str) -> str:
        v = getattr(obj, attr, None)
        return v.value if hasattr(v, "value") else (str(v) if v is not None else "")

    payload = {
        "meta": {
            "version": "1",
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "app": "nextfolio",
        },
        "accounts": [
            {
                "id": account_id_map[a.id],
                "name": a.name,
                "type": _val(a, "type"),
                "broker": a.broker or "",
                "currency": a.currency or "EUR",
                "url": a.url or "",
            }
            for a in accounts
        ],
        "assets": [
            {
                "id": asset_id_map[a_id],
                "symbol": asset.symbol,
                "name": asset.name,
                "isin": asset.isin or "",
                "type": _val(asset, "type"),
                "exchange": _val(asset, "exchange"),
                "currency": asset.currency,
                "sector": asset.sector or "",
            }
            for a_id, asset in assets_seen.items()
        ],
        "transactions": [
            {
                "account_id": account_id_map[tx.account_id],
                "asset_id": asset_id_map[tx.asset_id],
                "type": _val(tx, "type"),
                "date": tx.date.isoformat(),
                "quantity": float(tx.quantity),
                "price": float(tx.price),
                "fee": float(tx.fee),
                "price_currency": tx.price_currency,
                "exchange_rate": float(tx.exchange_rate),
                "fee_currency": tx.fee_currency,
                "notes": tx.notes or "",
            }
            for tx in transactions
            if tx.account_id in account_id_map and tx.asset_id in asset_id_map
        ],
    }

    filename = f"nextfolio_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    content = json.dumps(payload, ensure_ascii=False, indent=2)
    return StreamingResponse(
        iter([content]),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
